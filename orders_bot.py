from datetime import datetime, date, timedelta
from time import sleep
import os.path
import sys

import urllib.parse  # to decode from [%20Distressed%20] to [ Distressed ]
import requests
import openpyxl

from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox, QGroupBox, QLineEdit, QPushButton, \
    QGridLayout, QVBoxLayout, QWidget, QSizePolicy, QScrollArea, QSpacerItem, QSpinBox

from steampy.client import SteamClient
from steampy.models import Currency
from steampy.utils import GameOptions


def get_item_sales_for_week(steam_client, item_link, item_name):
    dict_item_average_sales = {}
    # getting average of item sales for the week
    dict_months = {
        'Jan': 1,
        'Feb': 2,
        'Mar': 3,
        'Apr': 4,
        'May': 5,
        'Jun': 6,
        'Jul': 7,
        'Aug': 8,
        'Sep': 9,
        'Oct': 10,
        'Nov': 11,
        'Dec': 12,
    }
    type_of_game = int(item_link.split('/')[5])  # 730 - cs:go game, 753 - steam card
    response_history_sold_items = []

    dict_number_game = {
        753: GameOptions.STEAM,
        730: GameOptions.CS,
        570: GameOptions.DOTA2,
        440: GameOptions.TF2,
        578080: GameOptions.PUBG,
        252490: GameOptions.RUST
    }

    try:
        print(f"Пытаемся взять историю продаж за неделю предмета {item_name}...")
        sleep(60)
        response_history_sold_items = steam_client.market.fetch_price_history(item_name, dict_number_game[type_of_game])
        print(f"Длина спарсенной истории предмета {item_name} = {len(response_history_sold_items)}")

        # we have history of ALL sales if this item
        # but we need info for last week:
        if response_history_sold_items:
            date_today = datetime.now().date()
            sum_count_sold_items = 0  # count sales of current item
            try:
                for buy_event in reversed(response_history_sold_items['prices']):
                    # from ['Jul 02 2014 01: +0', 417.777, '40'] to [Jul, 02, 2014]
                    date_buying_list = buy_event[0].split(':')[0].split(' ')[:-1]
                    date_of_buying = date(int(date_buying_list[2]), dict_months[date_buying_list[0]],
                                          int(date_buying_list[1]))
                    if date_of_buying > (date_today - timedelta(days=7)):  # info for only last week
                        sum_count_sold_items += int(buy_event[2])

                if int(sum_count_sold_items / 7) != 0:
                    dict_item_average_sales[item_name] = [int(sum_count_sold_items / 7)]
                    print(f"Предмет {item_name}. За неделю продано {sum_count_sold_items} штук. В среднем продаж за сутки = {sum_count_sold_items // 7}")
                    return sum_count_sold_items // 7

            except Exception as ex:
                print(f"{datetime.now().strftime('%H:%M:%S')}: Ошибка при вычислении продаж предмета {item_name}: {ex}")
                return 0

        else:
            print(f"{datetime.now().strftime('%H:%M:%S')}: Стим не выдал историю предмета {item_name}. Ждём 4 минуты")
            sleep(240)
            return 0

    except Exception as ex:
        print(f"{datetime.now().strftime('%H:%M:%S')}: Ошибка при получении истории продаж предмета {item_name}: {ex}, история: {response_history_sold_items}")
        return 0


def authorization_steam_account(login, password, maFile):
    """
    Authorization at steam account using login, password and his mafile
    Returns session of current account
    """
    steam_client = 0
    is_session_alive = False
    print(f"{datetime.now().strftime('%H:%M:%S')}: Пытаемся авторизоваться в аккаунт {login}")
    while not is_session_alive:
        try:
            steam_client = SteamClient('CDCE9BB5EE991C6B84CD9C9501C43A40')
            # login, password, C:\Users\me\Desktop\sda 1\maFiles\76561198877671188.maFile
            steam_client.login(login, password, maFile)
            # we have to delete code (code consists from 5 digits, which was sent to the phone) from steam
            # authenticator then maFile will be decoded
            is_session_alive = steam_client.is_session_alive()
        except Exception as ex:
            print(f"{datetime.now().strftime('%H:%M:%S')}: Ошибка авторизации {login}: {ex} попытка через 4 минуты:")
            sleep(240)
    print(f"{datetime.now().strftime('%H:%M:%S')}: Авторизация аккаунта {login} успешна!")
    return steam_client


def get_price_left_glass(average_sales_of_current_item, item_code, item_percent):
    try:
        # sum of walls in left glass should be in range of var=average_sales_of_current_item
        url_get_left_column = f"https://steamcommunity.com/market/itemordershistogram?country=RU&language=russian&currency=5&item_nameid={item_code}&two_factor=0"

        params = {
            'User-Agent': f"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36"
        }

        try:
            response = requests.get(url_get_left_column, params=params, timeout=30)
        except Exception as ex:
            print(f"Левый стакан: превышено время ожидания ответа сервера. Код предмета-{item_code}: {ex}")
            return 0

        buy_orders_100_json_left = response.json()
        print(f"Ответ сервера: {response.status_code}, длина левого стакана: {len(buy_orders_100_json_left)}")
        if buy_orders_100_json_left:
            # left glass
            matrix_last_100_sell_orders = buy_orders_100_json_left["sell_order_graph"]
            print(f"первые 10 позиций левого стакана: {matrix_last_100_sell_orders[:10]}")
            for order in matrix_last_100_sell_orders:
                count = order[1]
                if average_sales_of_current_item <= count:
                    price_left = float(order[0])
                    print(f"Найдена цена на {matrix_last_100_sell_orders.index(order) + 1} позиции = {price_left}")
                    print(order)
                    print(f"Вычитаем комиссию 13%:")
                    price_left *= 0.87
                    price_left = float(str(price_left)[:str(price_left).find('.') + 3])
                    print(f"{order[0]} - 13% = {price_left}")
                    print(f"вычитаем 1 копейку:")
                    before = price_left
                    price_left -= 0.01
                    price_left = round(price_left, 2)
                    print(f"{before} - 0.01 = {price_left}")
                    # minus percent from excel
                    print(f"Вычитаем {item_percent}%, которые указаны в таблице:")
                    before = price_left
                    price_left = price_left * ((100 - item_percent) / 100)
                    price_left = round(price_left, 2)
                    print(f"{before} - {item_percent}% = {price_left}")
                    return price_left, buy_orders_100_json_left
    except Exception as ex:
        print(f"{datetime.now().strftime('%H:%M:%S')}: Ошибка во время получения левого стакана: {ex}")
        return 0


def get_price_right_glass(average_sales_of_current_item, buy_orders_100_json):
    try:
        # sum of walls in right glass should be in range of var=average_sales_of_current_item
        # we got json of buy orders from function of left glass
        if buy_orders_100_json:
            # right glass
            matrix_last_100_buy_orders = buy_orders_100_json["buy_order_graph"]
            print(f"Первые 10 позиций правого стакана: {matrix_last_100_buy_orders[:10]}")
            for order in matrix_last_100_buy_orders:
                count = order[1]
                if average_sales_of_current_item <= count:
                    price_right = float(order[0])
                    print(f"Найдена цена на {matrix_last_100_buy_orders.index(order) + 1} позиции = {price_right}")
                    print(order)
                    print(f"Прибавляем 1 копейку:")
                    price_right += 0.01
                    price_right = round(price_right, 2)
                    print(f"{order[0]} + 0.01 = {float(price_right)}")
                    return price_right
    except Exception as ex:
        print(f"{datetime.now().strftime('%H:%M:%S')}: Ошибка во время обработки правого стакана: {ex}")
        return 0


def get_balace_of_account(steam_client, login):
    steam_wallet_balance = -2
    while steam_wallet_balance == -2:
        try:
            steam_wallet_balance = int(steam_client.get_wallet_balance()) - 1
            return steam_wallet_balance
        except Exception as ex:
            print(
                f"{datetime.now().strftime('%H:%M:%S')}: Ошибка при получении баланса аккаунта {login}: {ex}. Ждём 4 минуты")
            sleep(240)


def read_excel_get_item_history(steam_client, path_to_excel, login, dict_of_items_to_create_orders):
    """
    Read all items from Excel table
    Get dictionary of active order on current account
    Is there orders on current item in steam account
    If there is no active order for this item then get sales history
    """
    try:
        # read excel
        book = openpyxl.open(path_to_excel, data_only=True)
        sheet = book.active
        print(f"Количество строк в таблице {path_to_excel.split('/')[-1]}: {sheet.max_row}")
    except Exception as ex:
        print(
            f"{datetime.now().strftime('%H:%M:%S')}: Ошибка при открытии таблицы {path_to_excel.split('/')[-1]} аккаунта {login}: {ex}")
        return 0

    try:
        sleep(10)
        # getting dictionary of active orders and creating the same dictionary, where key = item name
        active_orders = steam_client.market.get_my_market_listings()['buy_orders']
        # key = id of creating buy order, but we need to make search by item name
    except Exception as ex:
        print(f"{datetime.now().strftime('%H:%M:%S')}: Ошибка при получении активных ордеров аккаунта {login}: {ex}")
        return 0

    try:
        dict_name_info = {}
        # also count amount of all active orders during the changing keys:
        sum_all_active_orders = 0
        for order_id, info in active_orders.items():
            dict_name_info[active_orders[order_id]['item_name']] = {
                'order_id': active_orders[order_id]['order_id'],
                'quantity': active_orders[order_id]['quantity'],
                'price': active_orders[order_id]['price']}
        print(f"Всего выставленных предметных ордеров: {len(dict_name_info)}")
        for item, info in dict_name_info.items():
            sum_all_active_orders += float(float(info['price'].split(' ')[0].replace(',', '.')) * int(info['quantity']))
        print(f"Сумма активных ордеров аккаунта {login} = {round(sum_all_active_orders) + 1}")
    except Exception as ex:
        print(f"{datetime.now().strftime('%H:%M:%S')}: Ошибка при подсчёте суммы активных ордеров аккаунта {login}: {ex}")
        return 0

    try:
        for row in range(2, sheet.max_row + 1):
            item_link = sheet[row][0].value  # Column A
            item_code = sheet[row][1].value  # Column B
            item_percent = sheet[row][2].value  # Column C

            if not item_link:  # empty line = finish of excel
                print(f"Встречена пустая строка №{row}")
                return dict_of_items_to_create_orders, round(sum_all_active_orders) + 1

            item_name = urllib.parse.unquote(item_link.split('/')[-1])
            print(f"{datetime.now().strftime('%H:%M:%S')}: Строка №{row}, предмет: {item_name}")
            type_of_game = int(item_link.split('/')[5])  # 730 or 753
            if type_of_game == 753:
                item_name_without_defis = '-'.join(item_name.split('-')[1:])
            else:
                item_name_without_defis = item_name

            if item_name_without_defis not in dict_name_info:  # if there is no active order for this item
                average_sales_of_current_item = get_item_sales_for_week(steam_client, item_link, item_name)
                if average_sales_of_current_item == 0:
                    print(f"Не удалось получить историю продаж за 7 дней, пропускаем предмет {item_name}")
                    print()
                    continue
                else:
                    sleep(10)
                    tuple_price_glasses = get_price_left_glass(average_sales_of_current_item, item_code, item_percent)
                    if isinstance(tuple_price_glasses, tuple):
                        price_left_glass, buy_orders_100_json = tuple_price_glasses[0], tuple_price_glasses[1]
                        print(f"Предмет {item_name} кол-во продаж за неделю = {average_sales_of_current_item} цена левого стакана = {price_left_glass}")
                        print(f"берем правый стакан:")
                        price_right_glass = get_price_right_glass(average_sales_of_current_item, buy_orders_100_json)
                        if price_right_glass == 0:
                            print(f"Не удалось получить правый стакан, пропускаем предмет {item_name}")
                            continue
                        else:
                            print(f"Предмет {item_name}. Продаж {average_sales_of_current_item}. Левый стакан {price_left_glass}. Правый стакан {price_right_glass}")
                            print(f"если цена правого стакана <= чем цена левого стакана, то сохраняем этот предмет для создания ордеров")
                            if price_right_glass <= price_left_glass:
                                dict_of_items_to_create_orders[item_name] = [average_sales_of_current_item,
                                                                             price_right_glass, type_of_game, item_code]
                                print(f"Добавили предмет {item_name}, продолжаем смотреть следующие предметы:")
                                print(dict_of_items_to_create_orders)
                            else:
                                print(f"Но данный момент ордер на предмет создавать не выгодно")
                            sleep(3)
                            print()
                    else:
                        if tuple_price_glasses is None:
                            print(f"Цена в левом стакане = None, пропускаем предмет {item_name} и ждём 4 минуты")
                            sleep(240)
                            continue
                        if tuple_price_glasses == 0:
                            print(f"Не удалось получить левый стакан, пропускаем предмет {item_name}, ждём 4 минуты")
                            sleep(240)
                            continue

            else:
                print(f"На предмет {item_name} уже есть выставленный ордер")

    except Exception as ex:
        print(f"{datetime.now().strftime('%H:%M:%S')}: Ошибка при чтении таблицы: {ex}")
        return 1


def write_history_table(path_to_excel, item_code_f, price):
    try:
        # read excel
        book = openpyxl.open(path_to_excel, data_only=True)
        sheet = book.active
        print(f"Открыли таблицу {path_to_excel} чтобы вставить цену ордера")

        for row in range(2, sheet.max_row + 1):
            item_link = sheet[row][0].value  # Column A
            item_code = sheet[row][1].value  # Column B

            if not item_link:
                break

            if str(item_code_f) in str(item_code):
                print(f"Нашли строчку с предметом {item_code}, смотрим ячейки таблицы:")
                # столбец E
                row_of_cur_history = sheet[row]
                miss_column = 0

                sum_for_aver_prices = 0
                count_prices = 1

                for cell in row_of_cur_history:
                    miss_column += 1
                    # находим среднее арифметическое
                    if miss_column >= 5 and cell.value and isinstance(cell.value, float):
                        sum_for_aver_prices += cell.value
                        count_prices += 1

                    if not cell.value and miss_column >= 5:
                        cell.value = price
                        sum_for_aver_prices += price
                        break
                sheet[row][3].value = round(sum_for_aver_prices / count_prices, 2)
        book.save(path_to_excel)
        print(f"Успешно записали цену и среднее арифметическое предмета в таблицу")
        return 1
    except Exception as ex:
        print(f"Ошибка при добавлении стоимости ордера: {ex}")
        return 0


def create_order_for_item(steam_client, item, price, count, type_of_game):
    # creates order
    sleep(5)

    dict_number_game = {
        753: GameOptions.STEAM,
        730: GameOptions.CS,
        570: GameOptions.DOTA2,
        440: GameOptions.TF2,
        578080: GameOptions.PUBG,
        252490: GameOptions.RUST
    }

    try:
        steam_client.market.create_buy_order(item, price, count, dict_number_game[type_of_game], Currency.RUB)
        return 1
    except Exception as ex:
        print(f"Не удалось создать ордер (либо он уже есть): {ex}")
        return 2


class Widget(QGroupBox):
    def __init__(self, name):
        super().__init__()
        self.setTitle(name.upper())
        self.name = name

        self.lineEditLogin = QLineEdit()  # поле ввода логина
        self.lineEditLogin.setPlaceholderText("логин")
        self.lineEditPassword = QLineEdit()  # поле ввода пароля
        self.lineEditPassword.setEchoMode(QLineEdit.Password)
        self.lineEditPassword.setPlaceholderText("пароль")

        self.btn_mafile = QPushButton(f'maFile {self.name}')  # текст на кнопке с выбором mafile
        self.btn_mafile.clicked.connect(self.open_mafile)
        self.lineEditPathMafile = QLineEdit()
        self.lineEditPathMafile.setPlaceholderText("путь до maFile")
        self.lineEditPathMafile.setReadOnly(True)

        self.btn_choose_excel = QPushButton(f'Excel {self.name}')  # текст на кнопке с выбором mafile
        self.btn_choose_excel.clicked.connect(self.open_excel)
        self.lineEditPathExcel = QLineEdit()
        self.lineEditPathExcel.setPlaceholderText("путь до таблицы Excel")
        self.lineEditPathExcel.setReadOnly(True)

        self.layout = QGridLayout(self)

        self.layout.addWidget(self.lineEditLogin, 0, 1)
        self.layout.addWidget(self.lineEditPassword, 1, 1)

        self.layout.addWidget(self.btn_mafile, 0, 2)
        self.layout.addWidget(self.lineEditPathMafile, 1, 2)

        self.layout.addWidget(self.btn_choose_excel, 0, 3)
        self.layout.addWidget(self.lineEditPathExcel, 1, 3)

    def open_mafile(self):
        """choose mafile of current account"""
        fname = QFileDialog.getOpenFileName(self, 'Выбрать maFile аккаунта', '')[0]
        self.lineEditPathMafile.setText(fname)

    def open_excel(self):
        fname = QFileDialog.getOpenFileName(self, 'Выбрать таблицу', '')[0]
        self.lineEditPathExcel.setText(fname)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Грамотное выставление ордеров steam')
        self.setGeometry(500, 300, 900, 450)

        centralWidget = QWidget()
        self.setCentralWidget(centralWidget)
        accounts_labels_list = [
            'Аккаунт №1', 'Аккаунт №2', 'Аккаунт №3', 'Аккаунт №4', 'Аккаунт №5', 'Аккаунт №6', 'Аккаунт №7',
            'Аккаунт №8', 'Аккаунт №9', 'Аккаунт №10', 'Аккаунт №11', 'Аккаунт №12', 'Аккаунт №13', 'Аккаунт №14',
            'Аккаунт №15', 'Аккаунт №16', 'Аккаунт №17', 'Аккаунт №18', 'Аккаунт №19', 'Аккаунт №20', 'Аккаунт №21',
            'Аккаунт №22', 'Аккаунт №23', 'Аккаунт №24', 'Аккаунт №25', 'Аккаунт №26', 'Аккаунт №27', 'Аккаунт №28',
            'Аккаунт №29', 'Аккаунт №30', 'Аккаунт №31', 'Аккаунт №32', 'Аккаунт №33', 'Аккаунт №34', 'Аккаунт №35',
            'Аккаунт №36', 'Аккаунт №37', 'Аккаунт №38', 'Аккаунт №39', 'Аккаунт №40', 'Аккаунт №41', 'Аккаунт №42',
            'Аккаунт №43', 'Аккаунт №44', 'Аккаунт №45', 'Аккаунт №46', 'Аккаунт №47', 'Аккаунт №48', 'Аккаунт №49',
            'Аккаунт №50', 'Аккаунт №51', 'Аккаунт №52', 'Аккаунт №53', 'Аккаунт №54', 'Аккаунт №55', 'Аккаунт №56',
            'Аккаунт №57', 'Аккаунт №58', 'Аккаунт №59', 'Аккаунт №60', 'Аккаунт №61', 'Аккаунт №62', 'Аккаунт №63',
            'Аккаунт №64', 'Аккаунт №65', 'Аккаунт №66', 'Аккаунт №67', 'Аккаунт №68', 'Аккаунт №69', 'Аккаунт №70',
            'Аккаунт №71', 'Аккаунт №72', 'Аккаунт №73', 'Аккаунт №74', 'Аккаунт №75', 'Аккаунт №76', 'Аккаунт №77',
            'Аккаунт №78', 'Аккаунт №79', 'Аккаунт №80', 'Аккаунт №81', 'Аккаунт №82', 'Аккаунт №83', 'Аккаунт №84',
            'Аккаунт №85', 'Аккаунт №86', 'Аккаунт №87', 'Аккаунт №88', 'Аккаунт №89', 'Аккаунт №90', 'Аккаунт №91',
            'Аккаунт №92', 'Аккаунт №93', 'Аккаунт №94', 'Аккаунт №95', 'Аккаунт №96', 'Аккаунт №97', 'Аккаунт №98',
            'Аккаунт №99', 'Аккаунт №100']

        self.controls = QWidget()
        self.controlsLayout = QVBoxLayout(self.controls)
        self.widgets = []
        right_index = 0

        for name in accounts_labels_list:
            try:
                file_path_data = 'data_login_pass.txt'
                if os.path.exists(file_path_data):
                    with open(file_path_data, 'r', encoding='utf-8') as file:
                        data = list(map(lambda x: x.rstrip(), file.readlines()))
                        item = Widget(name)
                        item.lineEditLogin.setText(data[right_index])
                        item.lineEditPassword.setText(data[right_index + 1])
                        item.lineEditPathMafile.setText(data[right_index + 2])
                        item.lineEditPathExcel.setText(data[right_index + 3])
                        right_index += 4
                        self.controlsLayout.addWidget(item)
                        self.widgets.append(item)
                else:
                    item = Widget(name)
                    self.controlsLayout.addWidget(item)
                    self.widgets.append(item)
            except IndexError:
                item = Widget(name)
                self.controlsLayout.addWidget(item)
                self.widgets.append(item)

        spacer = QSpacerItem(1, 1, QSizePolicy.Minimum, QSizePolicy.Expanding)
        self.controlsLayout.addItem(spacer)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(self.controls)

        self.spinBox = QSpinBox()
        self.spinBox.setRange(0, 100)
        self.spinBox.setProperty("value", 100)
        self.spinBox.valueChanged.connect(self.update_display)

        self.start_btn = QPushButton('Старт')  # Кнопка запуска бота
        self.start_btn.clicked.connect(self.run_bot)

        containerLayout = QVBoxLayout(centralWidget)
        containerLayout.addWidget(self.spinBox)
        containerLayout.addWidget(self.start_btn)
        containerLayout.addWidget(self.scroll)

    def run_bot(self):
        try:
            print(f"Количество аккаунтов: {self.spinBox.value()}")
            with open('data_login_pass.txt', 'w', encoding='utf-8') as file:
                for i in range(self.spinBox.value()):
                    account = self.widgets[i]

                    print(f"{account.lineEditLogin.text()}\n{account.lineEditPassword.text()}\n{account.lineEditPathMafile.text()}\n{account.lineEditPathExcel.text()}\n")
                    print(f"{account.lineEditLogin.text()}\n{account.lineEditPassword.text()}\n{account.lineEditPathMafile.text()}\n{account.lineEditPathExcel.text()}\n", file=file)

            while True:
                for i in range(self.spinBox.value()):
                    try:
                        account = self.widgets[i]

                        account_login = account.lineEditLogin.text()
                        account_password = account.lineEditPassword.text()
                        path_mafile = account.lineEditPathMafile.text()
                        path_excel_table = account.lineEditPathExcel.text()

                        # авторизация аккаунта
                        steam_client = authorization_steam_account(account_login, account_password, path_mafile)

                        dict_item_sales_price = {}
                        dict_item_average_sales_and_sum_orders = read_excel_get_item_history(steam_client,
                                                                                             path_excel_table,
                                                                                             account_login,
                                                                                             dict_item_sales_price)

                        if dict_item_average_sales_and_sum_orders == 1:
                            print(f"{datetime.now().strftime('%H:%M:%S')}: Из-за ошибки пропускаем предмет, ждём 4 минуты")
                            sleep(240)
                            continue
                        elif dict_item_average_sales_and_sum_orders == 0:
                            print(f"{datetime.now().strftime('%H:%M:%S')}: Из-за ошибки пропускаем аккаунт {account_login}, ждём 4 минуты")
                            sleep(240)
                            continue
                        else:
                            dict_item_sales_price, sum_all_active_orders = dict_item_average_sales_and_sum_orders[0], \
                                                                           dict_item_average_sales_and_sum_orders[1]
                            print(f'Словарь предметов для создания ордеров: {dict_item_sales_price}')
                            balance = get_balace_of_account(steam_client, account_login)
                            if balance == 0:
                                print(f"Не удалось получить баланс аккаунта {account_login}, либо на аккаунте 2 рубля -> пропускаем аккаунт")
                                sleep(60)
                                continue
                            else:
                                print(f"Баланс аккаунта {account_login} = {balance}")
                                print(f"Высчитываем по формуле кол-во ордеров для каждого предмета:")

                                # history of sales of ALL items
                                history_all_items = sum(info[0] for name, info in dict_item_sales_price.items())
                                print(f"История продаж всех предметов = {history_all_items}")

                                # balance * 10 - sum of active orders
                                access_balance = balance * 10 - sum_all_active_orders

                                for item_name, list_sales_price in sorted(dict_item_sales_price.items(),
                                                                          key=lambda x: x[1][0]):
                                    print()
                                    type_game = list_sales_price[2]
                                    item_code = list_sales_price[3]
                                    # current item sales history / history of sales of all items
                                    a = list_sales_price[0] / history_all_items

                                    # b = a / price of order
                                    b = a / list_sales_price[1]

                                    # c = b * (balace * 10 - sum of active orders)
                                    c = int(b * access_balance)

                                    price_one_order = float(list_sales_price[1])
                                    is_pretty_count = False
                                    # if price of order * c > then fact balance, we should decrease c
                                    while not is_pretty_count:
                                        if float(price_one_order * c) >= balance:
                                            c -= 1
                                        else:
                                            is_pretty_count = True
                                    # make order
                                    print(f"{datetime.now().strftime('%H:%M:%S')}: Выставляем ордера на предмет {item_name} в количестве {c} по цене {price_one_order}, цена в копейках = {round(price_one_order * 100)}")

                                    created_order = create_order_for_item(steam_client, item_name,
                                                                          f"{round(price_one_order * 100)}", c,
                                                                          type_game)
                                    if created_order == 0:
                                        print(f"{datetime.now().strftime('%H:%M:%S')}: Из-за ошибки создания ордера пропускаем предмет {item_name} и ждём 4 минуты")
                                        sleep(240)
                                        continue
                                    elif created_order == 2:
                                        print(f"{datetime.now().strftime('%H:%M:%S')}: Проверьте предмет {item_name}")
                                        sleep(5)
                                        continue
                                    elif created_order == 1:
                                        print(
                                            f"{datetime.now().strftime('%H:%M:%S')}: Успешно создан ордер на предмет {item_name}")
                                        access_balance = access_balance - int((price_one_order * c)) - 1
                                        print(f"Записываем цену в таблицу:")
                                        write_history_table(path_excel_table, item_code, price_one_order)
                    except requests.exceptions.ConnectionError as ex:
                        print(f"{datetime.now().strftime('%H:%M:%S')}: Во время работы с аккаунтом {self.widgets[i].lineEditLogin.text()} произошла ошибка с интернетом: {ex}")
                        print('Ждём 4 минуты и продолжим')
                        sleep(240)

        except Exception as ex:
            print(f"{datetime.now().strftime('%H:%M:%S')}: ВНИМАНИЕ! ГЛОБАЛЬНАЯ ОШИБКА В РАБОТЕ ПРОГРАММЫ: {ex}")

    def update_display(self, value):
        for i, widget in enumerate(self.widgets):
            if i + 1 <= value:
                widget.show()
            else:
                widget.hide()

    def sizeHint(self):
        return self.scroll.widget().size() * 0.7

    def closeEvent(self, event):
        try:  # is user sure to close application?
            reply = QMessageBox.question(
                self, 'Закрытие программы', 'Закрыть программу?',
                QMessageBox.Yes, QMessageBox.No)
            if reply == QMessageBox.Yes:
                event.accept()
            else:
                event.ignore()
        except Exception as ex:
            print(f"Ошибка при закрытии программы: {ex}")


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.excepthook = except_hook
    sys.exit(app.exec_())
